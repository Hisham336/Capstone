#excluding these drugs: Clozapine, Risperidone, Thioridazine 

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill

# Replace 'your_file.xlsx' with the path to your Excel file
file_path = 'drug_data.xlsx'

# Read the Excel file
data_frame = pd.read_excel(file_path)

#Finding all the wells
values_column = data_frame['Unnamed: 4'] 

# Converting the values column to a list
values_list = values_column.tolist()

wb = openpyxl.Workbook()

# unique_values = set(values_list)
genotypeStr = ""
for value in values_list:
    genotypeStr += ', ' + value.strip()

unique_values = set(genotypeStr.strip().split(','))

allGenotypes = []
for value in sorted(unique_values):
    value = value.strip()
    if(len(value) > 0 and value[0] == '*'):
        allGenotypes.append(value)

#drugPhenoAndGenotype is list of tuple (phenotype, [genotypes_asList])
#wb is excel workbook
#drugName is name of the drug on first column of the sheet
#allGenotype is a non-duplicate list of all genotypes on the wells
def fillSheet(genoTypesPos, drugPhenoAndGenotype, ws):
    phenotype, genotypeList = drugPhenoAndGenotype
    for genotype in genotypeList:
        column, row = genoTypesPos.get(genotype)
        if "Normal" in phenotype and "Intermediate" in phenotype:
            ws.cell(row=row, column=column).fill = PatternFill(start_color="5FBDD7EE", end_color="5FBDD7EE", fill_type="solid")
            ws.cell(row=1, column=16).value = phenotype
            ws.cell(row=1, column=16).fill = PatternFill(start_color="5FBDD7EE", end_color="5FBDD7EE", fill_type="solid")
        elif "Ultrarapid" in phenotype:
            ws.cell(row=row, column=column).fill = PatternFill(start_color="5FC9E6C7", end_color="5FC9E6C7", fill_type="solid")
            ws.cell(row=2, column=16).value = phenotype
            ws.cell(row=2, column=16).fill = PatternFill(start_color="5FC9E6C7", end_color="5FC9E6C7", fill_type="solid")
        elif "Normal" in phenotype:
            ws.cell(row=row, column=column).fill = PatternFill(start_color="5FFFFFCC", end_color="5FFFFFCC", fill_type="solid")
            ws.cell(row=3, column=16).value = phenotype
            ws.cell(row=3, column=16).fill = PatternFill(start_color="5FFFFFCC", end_color="5FFFFFCC", fill_type="solid")
        elif "Intermediate" in phenotype:
            ws.cell(row=row, column=column).fill = PatternFill(start_color="5FFFE5CC", end_color="5FFFE5CC", fill_type="solid")
            ws.cell(row=4, column=16).value = phenotype
            ws.cell(row=4, column=16).fill = PatternFill(start_color="5FFFE5CC", end_color="5FFFE5CC", fill_type="solid")
        elif "Poor" in phenotype:
            ws.cell(row=row, column=column).fill = PatternFill(start_color="5FFFD9EB", end_color="5FFFD9EB", fill_type="solid")
            ws.cell(row=5, column=16).value = phenotype
            ws.cell(row=5, column=16).fill = PatternFill(start_color="5FFFD9EB", end_color="5FFFD9EB", fill_type="solid")
        else:
            ws.cell(row=row, column=column).fill = PatternFill(start_color="5FE0E0E0", end_color="5FE0E0E0", fill_type="solid")
            ws.cell(row=6, column=16).value = phenotype
            ws.cell(row=6, column=16).fill = PatternFill(start_color="5FE0E0E0", end_color="5FE0E0E0", fill_type="solid")

def generate_sheet(wb, drugName, allGenotypes, drugPhenoAndGenotype):
    ws = wb.create_sheet(title=drugName)
    numOfWells = len(allGenotypes)
    i = 0
    wellMaxed = False
    genoTypesPos = {}
    ws.cell(row=1, column=1).value = "Row/Column"
    for row in range(2,10):
        ws.cell(row=row, column=1).value = row-1
    for column in range(2,14):
        ws.cell(row=1, column=column).value = column-1
    for row in range(2,10):
        if wellMaxed:
            break
        for column in range(2,14):
            if wellMaxed:
                break
            ws.cell(row=row, column=column).value = allGenotypes[i]
            genoTypesPos.update({allGenotypes[i] : (column, row)})
            i += 1
            if i == numOfWells-1:
                wellMaxed = True
            
    fillSheet(genoTypesPos, drugPhenoAndGenotype, ws)
    for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width
    return genoTypesPos, ws

genoTypesPos = None
ws = None
i = 0
for index, row in data_frame.iterrows():
    if i == 0:
        i = 1
        continue
    name = row['Unnamed: 0']
    phenotype = row['Unnamed: 2']
    genotypes = row["Unnamed: 4"]
    genotypeList = []
    for genotype in genotypes.split(","):
        genotype = genotype.strip()
        if(len(genotype) > 0 and genotype[0] == '*'):
            genotypeList.append(genotype)
    if(not pd.isnull(row['Unnamed: 0'])):
        genoTypesPos, ws = generate_sheet(wb, name, allGenotypes, (phenotype.strip(), genotypeList))
    else:
        fillSheet(genoTypesPos, (phenotype.strip(), genotypeList), ws)

        
# Remove the default sheet
wb.remove(wb["Sheet"])

# Save the workbook
wb.save("drug_genotypes.xlsx")